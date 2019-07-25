#kai excel orgaziner 0.0, data must be organized by set in txt
import openpyxl, logging, os
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s : %(message)s')

def sheet_char():
    """Requests and Returns number of Columns for dataset"""
    while True:
        columnnum = input('Input How Many Columns of Data as an Integer\n')
        try:
            int(columnnum)
            return columnnum
            break
        except:print('Not An Integer\n')

def solo_create(filename):
    """Creates a new Excel file with filename
       if file name is a blank string it uses Default_Excel.xlsx"""
    wb = openpyxl.Workbook()
    if filename == '': filename = 'Default_Excel.xlsx'
    filename = str(filename)
    wb.save(filename)
    
    return filename

def file_check(basename = 'default_excel_'):
    """Checks for a a file named 'default_excel_01.xlsx' then incriments
    the numbers until the file name does not locally exist.
    """
    numfile = 1
    while True:
        if numfile > 9:
            filename = basename + str(numfile)+'.xlsx'
            pass
        else:
            filename = basename + '0' + str(numfile) + '.xlsx'
            pass
        logging.debug(filename)
        if os.path.isfile(filename) == True:
            logging.debug('File exists')
            numfile += 1
            continue
        else:
            logging.debug('File does not exist')
            return filename
        break    
        break

def header_write(filename, col_count):
    """writes column discriptors in the top row of excel, left to right"""
    wb = openpyxl.load_workbook(filename)
    s01 = wb.active
    for count in range(1,int(col_count)+1):
        title = input('What would you like titled in column %d?\n' %count) 
        s01.cell(row=1, column=count).value = title
    wb.save(filename)

def column_write(dataname, filename,head, col_count):
    """Takes data by line and weather header exist and 
    number of columns and writes the data by line from dataname (.txt only)
    """
    colid = 1
    wb = openpyxl.load_workbook(filename)
    s01 = wb.active
    if head == True:
        rowid = 2
    else:
        rowid = 1
    fr = open(dataname, 'rb')
    for item in fr:
        if item.decode('utf-8', 'replace') == '\r\n':#
            continue
        s01.cell(row=rowid, column=colid).value = item.decode('utf-8', 'replace')
        """
            s01.cell(row=rowid, column=colid).value = 'error'
            print('Cell in Row %d Column %d has encountered a problem' %(rowid,colid))
        """
        colid += 1
        if colid > int(col_count): 
            colid = 1
            rowid += 1
    wb.save(filename)
    fr.close()

def codetest():
    """"runs a test of the methods"""
    coltest = sheet_char()
    nametest = file_check()
    solo_create(nametest)
    header_write(nametest, coltest)
    column_write('GDQ_donation_pull.txt', nametest, True, coltest)

codetest()
input ('End of Test')
